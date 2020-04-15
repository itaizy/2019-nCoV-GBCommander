import datetime
import json
import os

from openpyxl import Workbook

import pymysql

class updateStatistic():
    # MYSQL_HOST = '10.1.1.56'
    MYSQL_HOST = 'localhost'
    MYSQL_PORT = 3306
    MYSQL_USER = 'root'
    MYSQL_PASSWORD = "123456"
    MYSQL_DB = "ncov"

    def statistic_first(self, title, countries):
        count_tag = {
            "confirmedCount": "累计确诊",
            "confirmedIncr": "新增确诊",
            "curedCount": "累计治愈",
            "curedIncr": "新增治愈",
            "deadCount": "累计死亡",
            "deadIncr": "新增死亡",
        }
        with pymysql.connect(
                host=self.MYSQL_HOST,
                port=self.MYSQL_PORT,
                user=self.MYSQL_USER,
                passwd=self.MYSQL_PASSWORD,
                db=self.MYSQL_DB,
                charset='utf8mb4'
        ) as conn:
            wb = Workbook()
            for tag in count_tag:
                ws = wb.create_sheet(count_tag[tag])
                date_sql = "SELECT `date` FROM ncov_data_jhu WHERE `name` IN (" + ",".join(["'" + j + "'" for j in countries]) + ") GROUP BY `date`"
                conn.execute(date_sql)
                date_results = conn.fetchall()
                y = 2
                date_dict = {}
                ws.cell(1, 1).value = "国家名"
                for date in date_results:
                    ws.cell(row=1, column=y).value = date[0].strftime("%m/%d")
                    date_dict[date[0]] = y
                    y += 1
                data_sql = "SELECT `name`, `date`, `" + tag + "` FROM ncov_data_jhu WHERE `name` IN (" + ",".join(["'" + j + "'" for j in countries]) + ")"
                conn.execute(data_sql)
                count_dict = {}
                data_results = conn.fetchall()
                for data in data_results:
                    if data[0] not in count_dict.keys():
                        count_dict[data[0]] = {}
                    count_dict[data[0]][data[1]] = data[2]
                x = 2
                for country in count_dict:
                    ws.cell(x, 1).value = country
                    for date in count_dict[country]:
                        ws.cell(x, date_dict[date]).value = count_dict[country][date]
                    x += 1
            wb.remove(wb['Sheet'])
            wb.save(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "statistic", title + ".xlsx"))
            update_sql = "UPDATE statistic_info SET `updateTime` = '{}' WHERE `title` = '{}'"
            conn.execute(update_sql.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), title))

    def country_first(self, title, countries):
        with pymysql.connect(
                host=self.MYSQL_HOST,
                port=self.MYSQL_PORT,
                user=self.MYSQL_USER,
                passwd=self.MYSQL_PASSWORD,
                db=self.MYSQL_DB,
                charset='utf8mb4'
        ) as conn:
            wb = Workbook()
            temp = [title]
            temp.extend(countries)
            if title not in countries:
                countries = temp
            tag = ["日期", "累计确诊", "新增确诊", "累计治愈", "新增治愈", "累计死亡", "新增治愈", "死亡率"]
            for country in countries:
                ws = wb.create_sheet(country)
                date_sql = "SELECT `date` FROM ncov_data_jhu WHERE `name` IN (" + ",".join(["'" + j + "'" for j in countries]) + ") GROUP BY `date`"
                conn.execute(date_sql)
                date_results = conn.fetchall()
                date_dict = {}
                for i in range(len(tag)):
                    ws.cell(1, i+1).value = tag[i]
                x = 2
                for date in date_results:
                    ws.cell(row=x, column=1).value = date[0].strftime("%m/%d")
                    date_dict[date[0]] = x
                    x += 1
                if country != title:
                    data_sql = "SELECT `date`, `confirmedCount`, `confirmedIncr`, `curedCount`, `curedIncr`, `deadCount`, `deadIncr` " \
                               "FROM ncov_data_jhu WHERE `name` = '{}'".format(country)
                else:
                    data_sql = "SELECT `date`, SUM(`confirmedCount`), SUM(`confirmedIncr`), SUM(`curedCount`), SUM(`curedIncr`), " \
                               "SUM(`deadCount`), SUM(`deadIncr`) FROM ncov_data_jhu WHERE `name` IN (" + ",".join(["'" + j + "'" for j in countries]) + ") GROUP BY `date`"
                conn.execute(data_sql)
                data_results = conn.fetchall()
                for data in data_results:
                    ws.cell(date_dict[data[0]], 2).value = data[1]
                    ws.cell(date_dict[data[0]], 3).value = data[2]
                    ws.cell(date_dict[data[0]], 4).value = data[3]
                    ws.cell(date_dict[data[0]], 5).value = data[4]
                    ws.cell(date_dict[data[0]], 6).value = data[5]
                    ws.cell(date_dict[data[0]], 7).value = data[6]
                    ws.cell(date_dict[data[0]], 8).value = str(round(float(data[5]) * 100/(float(data[1]) if int(data[1]) != 0 else 1), 2)) + "%"
            wb.remove(wb['Sheet'])
            wb.save(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "statistic", title + ".xlsx"))
            update_sql = "UPDATE statistic_info SET `updateTime` = '{}' WHERE `title` = '{}'"
            conn.execute(update_sql.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), title))

    def update(self):
        with pymysql.connect(
                host=self.MYSQL_HOST,
                port=self.MYSQL_PORT,
                user=self.MYSQL_USER,
                passwd=self.MYSQL_PASSWORD,
                db=self.MYSQL_DB,
                charset='utf8mb4'
        ) as conn:
            sql = "SELECT `title`, `countries`, `type`, `link` FROM `statistic_info`"
            conn.execute(sql)
            results = conn.fetchall()
            for result in results:
                if hasattr(self, result[2]):
                    f = getattr(self, result[2])
                    f(result[0], json.loads(result[1]))

if __name__ == '__main__':
    us = updateStatistic()
    us.update()